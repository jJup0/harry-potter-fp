#!/usr/bin/env python3
"""
Process Aitor's input data (v2):
1. Screen time xlsx -> data/metrics/screen_time_v2.json
2. Book mentions xlsx -> data/metrics/book_mentions_v2.json
3. FP Rules PDF -> data/fp_rules.txt (extracted text)
4. Screenplay PDFs -> data/raw/screenplays_v2/
5. Books epub -> data/raw/books_v2/ (if extractable)
"""
import json
import os
import re

import openpyxl
import fitz  # pymupdf

PROJECT_ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..")
INPUT_DIR = os.path.join(PROJECT_ROOT, "data", "freind-input-data")
METRICS_DIR = os.path.join(PROJECT_ROOT, "data", "metrics")
RAW_DIR = os.path.join(PROJECT_ROOT, "data", "raw")

FILM_MAP = {
    "Harry Potter and the Sorcerer's Stone": "1_philosophers_stone",
    "Harry Potter and the Chamber of Secrets": "2_chamber_of_secrets",
    "Harry Potter and the Prisoner of Azkaban": "3_prisoner_of_azkaban",
    "Harry Potter and the Goblet of Fire": "4_goblet_of_fire",
    "Harry Potter and the Order of the Phoenix": "5_order_of_the_phoenix",
    "Harry Potter and the Half-Blood Prince": "6_half_blood_prince",
    "Harry Potter and the Deathly Hallows: Part 1": "7_deathly_hallows_p1",
    "Harry Potter and the Deathly Hallows: Part 2": "8_deathly_hallows_p2",
}


def parse_screen_time_value(val):
    """Convert screen time value to minutes (float). Handles datetime.time and plain numbers."""
    import datetime
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, datetime.time):
        return val.hour + val.minute / 60.0
    return 0.0


def process_screen_time():
    """Process screen-time.xlsx -> screen_time_v2.json"""
    print("Processing screen time...")
    path = os.path.join(INPUT_DIR, "Screen Time (Movies)", "screen-time.xlsx")
    wb = openpyxl.load_workbook(path)

    # Use Sheet1 which has clean numeric minutes
    ws = wb["Sheet1"]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        movie, character, minutes, max_minutes = row
        if not character or not movie:
            continue
        character = character.strip()
        minutes = float(minutes) if minutes else 0

        film_key = None
        for full_name, key in FILM_MAP.items():
            if full_name in str(movie):
                film_key = key
                break
        if not film_key:
            continue

        if character not in result:
            result[character] = {}
        result[character][film_key] = minutes

    # Add totals
    for char in result:
        result[char]['_total'] = sum(v for k, v in result[char].items() if k != '_total')

    # Also process per-film sheets (2-7.2) for characters not in Sheet1
    for sheet_name in wb.sheetnames:
        if sheet_name == "Sheet1":
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 3:
                continue
            movie, character, screen_time = row[0], row[1], row[2]
            if not character or not movie:
                continue
            character = character.strip().lstrip("Professor ").strip()
            minutes = parse_screen_time_value(screen_time)
            if minutes == 0:
                continue

            movie_clean = str(movie).replace('\xa0', '').strip()
            film_key = None
            for full_name, key in FILM_MAP.items():
                if full_name in movie_clean:
                    film_key = key
                    break
            if not film_key:
                continue

            if character not in result:
                result[character] = {}
            if film_key not in result[character]:
                result[character][film_key] = minutes

    # Recalculate totals
    for char in result:
        result[char]['_total'] = round(sum(v for k, v in result[char].items() if k != '_total'), 2)

    result = dict(sorted(result.items(), key=lambda x: x[1].get('_total', 0), reverse=True))

    out_path = os.path.join(METRICS_DIR, "screen_time_v2.json")
    with open(out_path, 'w') as f:
        json.dump(result, f, indent=2)
    print(f"  {len(result)} characters -> {out_path}")
    for name, data in list(result.items())[:10]:
        print(f"    {name}: {data['_total']} min")


def process_book_mentions():
    """Process Harry Potter Mentions.xlsx -> book_mentions_v2.json"""
    print("\nProcessing book mentions...")
    path = os.path.join(INPUT_DIR, "Book Mentions", "Harry Potter Mentions.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]

    # Header row has book names
    headers = [cell.value for cell in ws[1]]
    book_keys = [
        "1_philosophers_stone", "2_chamber_of_secrets", "3_prisoner_of_azkaban",
        "4_goblet_of_fire", "5_order_of_the_phoenix", "6_half_blood_prince",
        "7_deathly_hallows"
    ]

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        character = vals[0]
        if not character:
            continue
        character = str(character).strip()

        char_data = {}
        for i, book_key in enumerate(book_keys):
            val = vals[i + 1] if i + 1 < len(vals) else None
            if isinstance(val, (int, float)) and val > 0:
                char_data[book_key] = int(val)

        if char_data:
            char_data['_total'] = sum(char_data.values())
            result[character] = char_data

    result = dict(sorted(result.items(), key=lambda x: x[1].get('_total', 0), reverse=True))

    out_path = os.path.join(METRICS_DIR, "book_mentions_v2.json")
    with open(out_path, 'w') as f:
        json.dump(result, f, indent=2)
    print(f"  {len(result)} characters -> {out_path}")
    for name, data in list(result.items())[:10]:
        print(f"    {name}: {data['_total']} mentions")


def process_fp_rules():
    """Extract FP rules PDF to text."""
    print("\nProcessing FP rules...")
    path = os.path.join(INPUT_DIR, "Reglas_de_FP_.pdf")
    doc = fitz.open(path)
    text = ""
    for page in doc:
        text += page.get_text() + "\n\n"

    out_path = os.path.join(PROJECT_ROOT, "data", "fp_rules.txt")
    with open(out_path, 'w') as f:
        f.write(text)
    print(f"  {len(doc)} pages, {len(text)} chars -> {out_path}")


def process_screenplay_pdfs():
    """Extract screenplay PDFs to text."""
    print("\nProcessing screenplay PDFs...")
    scripts_dir = os.path.join(INPUT_DIR, "Scripts Movies")
    out_dir = os.path.join(RAW_DIR, "screenplays_v2")
    os.makedirs(out_dir, exist_ok=True)

    pdf_map = {
        "HP1_Script.pdf": "1_philosophers_stone",
        "HP2_Script.pdf": "2_chamber_of_secrets",
        "HP3_Script.pdf": "3_prisoner_of_azkaban",
        "HP4_Script.pdf": "4_goblet_of_fire",
        "HP5_Script.pdf": "5_order_of_the_phoenix",
        "HP6_Script.pdf": "6_half_blood_prince",
        "HP7_1_Script.pdf": "7_deathly_hallows_p1",
        "HP7_2_Script-1.pdf": "8_deathly_hallows_p2",
    }

    for pdf_name, out_key in pdf_map.items():
        pdf_path = os.path.join(scripts_dir, pdf_name)
        if not os.path.exists(pdf_path):
            print(f"  SKIP {pdf_name} (not found)")
            continue
        doc = fitz.open(pdf_path)
        text = ""
        for page in doc:
            text += page.get_text() + "\n"

        out_path = os.path.join(out_dir, f"{out_key}.txt")
        with open(out_path, 'w') as f:
            f.write(text)
        print(f"  {pdf_name}: {len(doc)} pages, {len(text)} chars -> {out_key}.txt")


def process_books():
    """Extract books from epub/pdf."""
    print("\nProcessing books...")
    books_dir = os.path.join(INPUT_DIR, "Books")
    out_dir = os.path.join(RAW_DIR, "books_v2")
    os.makedirs(out_dir, exist_ok=True)

    # Try the PDF first (single file with all books)
    pdf_path = os.path.join(books_dir, "harrypotter-16.pdf")
    if os.path.exists(pdf_path):
        doc = fitz.open(pdf_path)
        text = ""
        for page in doc:
            text += page.get_text() + "\n"

        out_path = os.path.join(out_dir, "all_books.txt")
        with open(out_path, 'w') as f:
            f.write(text)
        print(f"  PDF: {len(doc)} pages, {len(text)} chars -> all_books.txt")

        # Try to split by book
        book_markers = [
            ("1_philosophers_stone", r"(?:Philosopher|Sorcerer)'s Stone"),
            ("2_chamber_of_secrets", r"Chamber of Secrets"),
            ("3_prisoner_of_azkaban", r"Prisoner of Azkaban"),
            ("4_goblet_of_fire", r"Goblet of Fire"),
            ("5_order_of_the_phoenix", r"Order of the Phoenix"),
            ("6_half_blood_prince", r"Half.Blood Prince"),
            ("7_deathly_hallows", r"Deathly Hallows"),
        ]
        # Find split points
        splits = []
        for key, pattern in book_markers:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                splits.append((key, match.start()))
        splits.sort(key=lambda x: x[1])

        if len(splits) >= 2:
            for i, (key, start) in enumerate(splits):
                end = splits[i + 1][1] if i + 1 < len(splits) else len(text)
                book_text = text[start:end]
                book_path = os.path.join(out_dir, f"{key}.txt")
                with open(book_path, 'w') as f:
                    f.write(book_text)
                print(f"    Split: {key} ({len(book_text)} chars)")

    # Also try epub
    epub_path = os.path.join(books_dir, "Harry Potter. La coleccion comp - J. K. Rowling(1).epub")
    if os.path.exists(epub_path):
        try:
            import ebooklib
            from ebooklib import epub
            from html.parser import HTMLParser

            class TextExtractor(HTMLParser):
                def __init__(self):
                    super().__init__()
                    self.text = []
                    self.skip = 0
                def handle_starttag(self, tag, attrs):
                    if tag in ('script', 'style'): self.skip += 1
                def handle_endtag(self, tag):
                    if tag in ('script', 'style') and self.skip > 0: self.skip -= 1
                    if tag in ('p', 'br', 'div', 'h1', 'h2', 'h3'): self.text.append('\n')
                def handle_data(self, data):
                    if self.skip == 0: self.text.append(data)
                def get_text(self):
                    return ''.join(self.text)

            book = epub.read_epub(epub_path)
            full_text = ""
            for item in book.get_items_of_type(ebooklib.ITEM_DOCUMENT):
                html = item.get_content().decode('utf-8', errors='replace')
                parser = TextExtractor()
                parser.feed(html)
                full_text += parser.get_text() + "\n"

            epub_out = os.path.join(out_dir, "all_books_epub.txt")
            with open(epub_out, 'w') as f:
                f.write(full_text)
            print(f"  EPUB: {len(full_text)} chars -> all_books_epub.txt")
            print(f"    NOTE: This is the SPANISH edition ('La colección completa')")
        except Exception as e:
            print(f"  EPUB extraction failed: {e}")


if __name__ == '__main__':
    process_screen_time()
    process_book_mentions()
    process_fp_rules()
    process_screenplay_pdfs()
    process_books()
    print("\nDone!")
